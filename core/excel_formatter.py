"""
Module de formatage Excel avancé pour Open Pandas-AI.
Utilise openpyxl pour appliquer des styles, conditional formatting, et mise en forme.
"""

from pathlib import Path
from typing import Dict, List, Optional, Any, Union, Tuple
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, 
    NamedStyle, Color
)
from openpyxl.formatting.rule import (
    ColorScaleRule, FormulaRule, CellIsRule,
    DataBarRule, IconSetRule
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook


# ============ STYLES PRÉDÉFINIS ============

# Couleurs
COLORS = {
    'header_bg': 'E0E0E0',
    'header_bg_dark': '4A4A4A',
    'success': '28A745',
    'warning': 'FFC107',
    'danger': 'DC3545',
    'info': '17A2B8',
    'light_gray': 'F5F5F5',
    'white': 'FFFFFF',
    'black': '000000'
}

# Bordures
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

MEDIUM_BORDER = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)


def apply_auto_column_width(worksheet: Worksheet, min_width: int = 8, max_width: int = 50):
    """
    Ajuste automatiquement la largeur des colonnes en fonction du contenu.
    
    Args:
        worksheet: Feuille openpyxl
        min_width: Largeur minimale
        max_width: Largeur maximale
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = max(min_width, min(max_length + 2, max_width))
        worksheet.column_dimensions[column_letter].width = adjusted_width


def apply_header_style(
    worksheet: Worksheet,
    header_row: int = 1,
    font_bold: bool = True,
    font_color: str = 'black',
    bg_color: str = 'header_bg',
    alignment: str = 'center',
    freeze: bool = True
):
    """
    Applique un style aux en-têtes de colonnes.
    
    Args:
        worksheet: Feuille openpyxl
        header_row: Numéro de la ligne d'en-tête (1-based)
        font_bold: Police en gras
        font_color: Couleur de la police
        bg_color: Couleur de fond (clé dans COLORS ou code hex)
        alignment: Alignement ('left', 'center', 'right')
        freeze: Geler la ligne d'en-tête
    """
    # Résoudre les couleurs
    bg_color_hex = COLORS.get(bg_color, bg_color)
    font_color_hex = COLORS.get(font_color, font_color)
    
    header_font = Font(bold=font_bold, color=font_color_hex)
    header_fill = PatternFill(start_color=bg_color_hex, end_color=bg_color_hex, fill_type="solid")
    header_alignment = Alignment(horizontal=alignment, vertical="center")
    
    for cell in worksheet[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = THIN_BORDER
    
    # Geler la ligne d'en-tête
    if freeze:
        worksheet.freeze_panes = f'A{header_row + 1}'


def apply_alternating_row_colors(
    worksheet: Worksheet,
    start_row: int = 2,
    odd_color: str = 'white',
    even_color: str = 'light_gray'
):
    """
    Applique une alternance de couleurs aux lignes pour une meilleure lisibilité.
    
    Args:
        worksheet: Feuille openpyxl
        start_row: Première ligne de données (1-based)
        odd_color: Couleur des lignes impaires
        even_color: Couleur des lignes paires
    """
    odd_fill = PatternFill(start_color=COLORS.get(odd_color, odd_color), 
                           end_color=COLORS.get(odd_color, odd_color), 
                           fill_type="solid")
    even_fill = PatternFill(start_color=COLORS.get(even_color, even_color), 
                            end_color=COLORS.get(even_color, even_color), 
                            fill_type="solid")
    
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=start_row), start=start_row):
        fill = even_fill if row_idx % 2 == 0 else odd_fill
        for cell in row:
            cell.fill = fill


def apply_conditional_formatting(
    worksheet: Worksheet,
    column_letter: str,
    rule_type: str = 'data_bar',
    start_row: int = 2,
    end_row: Optional[int] = None,
    **kwargs
):
    """
    Applique un formatage conditionnel à une colonne.
    
    Args:
        worksheet: Feuille openpyxl
        column_letter: Lettre de la colonne (ex: 'B', 'C')
        rule_type: Type de règle ('data_bar', 'color_scale', 'icon_set', 'cell_is')
        start_row: Première ligne de données
        end_row: Dernière ligne (None = dernière ligne avec données)
        **kwargs: Arguments spécifiques au type de règle
    """
    if end_row is None:
        end_row = worksheet.max_row
    
    cell_range = f"{column_letter}{start_row}:{column_letter}{end_row}"
    
    if rule_type == 'data_bar':
        # Barre de données
        color = kwargs.get('color', '638EC6')
        rule = DataBarRule(
            start_type="min", end_type="max",
            color=color
        )
        worksheet.conditional_formatting.add(cell_range, rule)
    
    elif rule_type == 'color_scale':
        # Échelle de couleurs (vert-jaune-rouge ou personnalisée)
        start_color = kwargs.get('start_color', '00FF00')  # Vert
        mid_color = kwargs.get('mid_color', 'FFFF00')  # Jaune
        end_color = kwargs.get('end_color', 'FF0000')  # Rouge
        
        rule = ColorScaleRule(
            start_type="min", start_color=start_color,
            mid_type="percentile", mid_value=50, mid_color=mid_color,
            end_type="max", end_color=end_color
        )
        worksheet.conditional_formatting.add(cell_range, rule)
    
    elif rule_type == 'icon_set':
        # Jeu d'icônes
        icon_style = kwargs.get('icon_style', '3Arrows')
        rule = IconSetRule(
            icon_style=icon_style,
            type="num",
            values=[0, 33, 67]
        )
        worksheet.conditional_formatting.add(cell_range, rule)
    
    elif rule_type == 'cell_is':
        # Règle basée sur valeur
        operator = kwargs.get('operator', 'greaterThan')
        formula = kwargs.get('formula', ['0'])
        fill_color = kwargs.get('fill_color', 'success')
        
        fill = PatternFill(
            start_color=COLORS.get(fill_color, fill_color),
            end_color=COLORS.get(fill_color, fill_color),
            fill_type="solid"
        )
        
        rule = CellIsRule(
            operator=operator,
            formula=formula,
            fill=fill
        )
        worksheet.conditional_formatting.add(cell_range, rule)


def apply_number_format(
    worksheet: Worksheet,
    column_letter: str,
    format_code: str,
    start_row: int = 2,
    end_row: Optional[int] = None
):
    """
    Applique un format de nombre à une colonne.
    
    Args:
        worksheet: Feuille openpyxl
        column_letter: Lettre de la colonne
        format_code: Code de format Excel (ex: '#,##0.00', '0.00%', 'YYYY-MM-DD')
        start_row: Première ligne
        end_row: Dernière ligne
    """
    if end_row is None:
        end_row = worksheet.max_row
    
    for row in range(start_row, end_row + 1):
        cell = worksheet[f"{column_letter}{row}"]
        cell.number_format = format_code


def apply_borders(
    worksheet: Worksheet,
    border_style: str = 'thin',
    start_row: int = 1,
    end_row: Optional[int] = None,
    start_col: int = 1,
    end_col: Optional[int] = None
):
    """
    Applique des bordures à une plage de cellules.
    
    Args:
        worksheet: Feuille openpyxl
        border_style: Style de bordure ('thin', 'medium', 'thick')
        start_row: Première ligne
        end_row: Dernière ligne
        start_col: Première colonne
        end_col: Dernière colonne
    """
    if end_row is None:
        end_row = worksheet.max_row
    if end_col is None:
        end_col = worksheet.max_column
    
    border = Border(
        left=Side(style=border_style),
        right=Side(style=border_style),
        top=Side(style=border_style),
        bottom=Side(style=border_style)
    )
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            worksheet.cell(row=row, column=col).border = border


def format_excel_output(
    file_path: Union[str, Path],
    format_options: Optional[Dict[str, Any]] = None
) -> str:
    """
    Applique un formatage complet à un fichier Excel existant.
    
    Args:
        file_path: Chemin du fichier Excel
        format_options: Options de formatage :
            - 'auto_width': bool (ajustement automatique largeur colonnes)
            - 'header_style': dict (options pour apply_header_style)
            - 'alternating_rows': bool (alternance couleurs lignes)
            - 'conditional_formatting': list of dict (règles de formatage conditionnel)
            - 'number_formats': dict {column: format_code}
            - 'freeze_panes': str (ex: 'A2')
            - 'borders': str (style de bordures)
    
    Returns:
        Chemin du fichier formaté
    """
    file_path = Path(file_path)
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Options par défaut
    if format_options is None:
        format_options = {
            'auto_width': True,
            'header_style': {},
            'alternating_rows': True,
            'borders': 'thin'
        }
    
    # Ajustement automatique des largeurs
    if format_options.get('auto_width', True):
        apply_auto_column_width(ws)
    
    # Style des en-têtes
    if 'header_style' in format_options:
        apply_header_style(ws, **format_options.get('header_style', {}))
    else:
        apply_header_style(ws)  # Valeurs par défaut
    
    # Alternance des couleurs de lignes
    if format_options.get('alternating_rows', False):
        apply_alternating_row_colors(ws)
    
    # Formatage conditionnel
    for cf_rule in format_options.get('conditional_formatting', []):
        column = cf_rule.pop('column', 'B')
        rule_type = cf_rule.pop('rule_type', 'data_bar')
        apply_conditional_formatting(ws, column, rule_type, **cf_rule)
    
    # Formats de nombres
    for column, fmt in format_options.get('number_formats', {}).items():
        apply_number_format(ws, column, fmt)
    
    # Gel des volets
    if 'freeze_panes' in format_options:
        ws.freeze_panes = format_options['freeze_panes']
    
    # Bordures
    if 'borders' in format_options:
        apply_borders(ws, border_style=format_options['borders'])
    
    wb.save(file_path)
    return str(file_path)


def format_excel_buffer(
    buffer: BytesIO,
    format_options: Optional[Dict[str, Any]] = None
) -> BytesIO:
    """
    Applique un formatage à un buffer Excel en mémoire.
    
    Args:
        buffer: Buffer contenant le fichier Excel
        format_options: Options de formatage (voir format_excel_output)
    
    Returns:
        Nouveau buffer avec le fichier formaté
    """
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb.active
    
    # Options par défaut
    if format_options is None:
        format_options = {
            'auto_width': True,
            'header_style': {},
            'borders': 'thin'
        }
    
    # Ajustement automatique des largeurs
    if format_options.get('auto_width', True):
        apply_auto_column_width(ws)
    
    # Style des en-têtes
    if 'header_style' in format_options or format_options.get('auto_width'):
        apply_header_style(ws, **format_options.get('header_style', {}))
    
    # Alternance des couleurs de lignes
    if format_options.get('alternating_rows', False):
        apply_alternating_row_colors(ws)
    
    # Formatage conditionnel
    for cf_rule in format_options.get('conditional_formatting', []):
        column = cf_rule.pop('column', 'B')
        rule_type = cf_rule.pop('rule_type', 'data_bar')
        apply_conditional_formatting(ws, column, rule_type, **cf_rule)
    
    # Formats de nombres
    for column, fmt in format_options.get('number_formats', {}).items():
        apply_number_format(ws, column, fmt)
    
    # Bordures
    if 'borders' in format_options:
        apply_borders(ws, border_style=format_options['borders'])
    
    # Sauvegarder dans un nouveau buffer
    output_buffer = BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    
    return output_buffer


# ============ STYLES PRÉDÉFINIS POUR RAPPORTS ============

def apply_report_style(worksheet: Worksheet, report_type: str = 'professional'):
    """
    Applique un style de rapport prédéfini.
    
    Args:
        worksheet: Feuille openpyxl
        report_type: Type de rapport ('professional', 'modern', 'minimal')
    """
    if report_type == 'professional':
        apply_auto_column_width(worksheet)
        apply_header_style(worksheet, bg_color='4A4A4A', font_color='white')
        apply_alternating_row_colors(worksheet)
        apply_borders(worksheet, border_style='thin')
    
    elif report_type == 'modern':
        apply_auto_column_width(worksheet)
        apply_header_style(worksheet, bg_color='17A2B8', font_color='white')
        apply_borders(worksheet, border_style='thin')
    
    elif report_type == 'minimal':
        apply_auto_column_width(worksheet)
        apply_header_style(worksheet, bg_color='F5F5F5', font_color='black', freeze=True)


def get_available_styles() -> List[str]:
    """Retourne la liste des styles de rapport disponibles."""
    return ['professional', 'modern', 'minimal']


def get_available_formats() -> Dict[str, str]:
    """Retourne les formats de nombres disponibles."""
    return {
        'currency_eur': '#,##0.00 €',
        'currency_usd': '$#,##0.00',
        'percentage': '0.00%',
        'number': '#,##0',
        'decimal_2': '#,##0.00',
        'date_iso': 'YYYY-MM-DD',
        'date_fr': 'DD/MM/YYYY',
        'datetime': 'YYYY-MM-DD HH:MM:SS'
    }

