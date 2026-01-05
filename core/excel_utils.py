"""
Module utilitaires Excel pour Open Pandas-AI.
Gère la lecture multi-sheets, l'export sécurisé, les pivot tables et le merging.
"""

import os
import tempfile
from pathlib import Path
from typing import Dict, List, Optional, Union, Any
from io import BytesIO

import pandas as pd

# Limites de sécurité
MAX_FILE_SIZE_MB = int(os.getenv('EXCEL_MAX_FILE_SIZE_MB', '50'))
MAX_SHEETS_LOADED = int(os.getenv('EXCEL_MAX_SHEETS_LOADED', '10'))


def detect_excel_sheets(file_path_or_buffer: Union[str, Path, BytesIO]) -> List[str]:
    """
    Détecte et retourne la liste des feuilles d'un fichier Excel.
    
    Args:
        file_path_or_buffer: Chemin du fichier ou buffer (UploadedFile de Streamlit)
    
    Returns:
        Liste des noms de feuilles
    """
    try:
        excel_file = pd.ExcelFile(file_path_or_buffer)
        return excel_file.sheet_names
    except Exception as e:
        raise ValueError(f"Impossible de lire les feuilles Excel: {e}")


def read_excel_multi_sheets(
    file_path_or_buffer: Union[str, Path, BytesIO],
    sheet_name: Optional[Union[str, int, List[str]]] = None
) -> Union[pd.DataFrame, Dict[str, pd.DataFrame]]:
    """
    Lit un fichier Excel avec support multi-sheets.
    
    Args:
        file_path_or_buffer: Chemin du fichier ou buffer
        sheet_name: 
            - None: retourne toutes les feuilles (dict)
            - str/int: retourne la feuille spécifiée (DataFrame)
            - List[str]: retourne les feuilles spécifiées (dict)
    
    Returns:
        DataFrame unique ou dictionnaire {nom_feuille: DataFrame}
    """
    try:
        if sheet_name is None:
            # Charger toutes les feuilles (avec limite de sécurité)
            sheets = detect_excel_sheets(file_path_or_buffer)
            if len(sheets) > MAX_SHEETS_LOADED:
                sheets = sheets[:MAX_SHEETS_LOADED]
            
            result = {}
            for sheet in sheets:
                # Reset buffer position if it's a BytesIO
                if hasattr(file_path_or_buffer, 'seek'):
                    file_path_or_buffer.seek(0)
                result[sheet] = pd.read_excel(file_path_or_buffer, sheet_name=sheet)
            return result
        else:
            # Reset buffer position if it's a BytesIO
            if hasattr(file_path_or_buffer, 'seek'):
                file_path_or_buffer.seek(0)
            return pd.read_excel(file_path_or_buffer, sheet_name=sheet_name)
    except Exception as e:
        raise ValueError(f"Erreur lors de la lecture Excel: {e}")


def export_dataframe_to_excel(
    df: pd.DataFrame,
    output_path: Union[str, Path],
    sheet_name: str = "Sheet1",
    format_options: Optional[Dict[str, Any]] = None,
    auto_format: bool = True
) -> str:
    """
    Exporte un DataFrame vers un fichier Excel avec formatage optionnel.
    
    Args:
        df: DataFrame à exporter
        output_path: Chemin de sortie
        sheet_name: Nom de la feuille
        format_options: Options de formatage (voir excel_formatter.py)
        auto_format: Active le formatage automatique (largeurs colonnes, en-têtes)
    
    Returns:
        Chemin du fichier créé
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    output_path = Path(output_path)
    
    # Export initial avec pandas
    df.to_excel(output_path, sheet_name=sheet_name, index=False, engine='openpyxl')
    
    if auto_format or format_options:
        # Appliquer le formatage
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Auto-ajustement des largeurs de colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max 50 caractères
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Style des en-têtes
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Appliquer options personnalisées si fournies
        if format_options:
            _apply_custom_format_options(ws, format_options)
        
        wb.save(output_path)
    
    return str(output_path)


def _apply_custom_format_options(worksheet, format_options: Dict[str, Any]):
    """
    Applique des options de formatage personnalisées.
    
    Args:
        worksheet: Feuille openpyxl
        format_options: Dict avec les options
            - 'freeze_panes': str (ex: 'A2' pour geler la première ligne)
            - 'number_format': dict {column: format}
            - 'column_widths': dict {column: width}
    """
    from openpyxl.styles import numbers
    
    if 'freeze_panes' in format_options:
        worksheet.freeze_panes = format_options['freeze_panes']
    
    if 'number_format' in format_options:
        for col, fmt in format_options['number_format'].items():
            for cell in worksheet[col]:
                cell.number_format = fmt
    
    if 'column_widths' in format_options:
        for col, width in format_options['column_widths'].items():
            worksheet.column_dimensions[col].width = width


def export_dataframe_to_buffer(
    df: pd.DataFrame,
    sheet_name: str = "Sheet1",
    auto_format: bool = True
) -> BytesIO:
    """
    Exporte un DataFrame vers un buffer BytesIO (pour téléchargement Streamlit).
    
    Args:
        df: DataFrame à exporter
        sheet_name: Nom de la feuille
        auto_format: Active le formatage automatique
    
    Returns:
        BytesIO contenant le fichier Excel
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    buffer = BytesIO()
    
    # Export initial
    df.to_excel(buffer, sheet_name=sheet_name, index=False, engine='openpyxl')
    
    if auto_format:
        buffer.seek(0)
        wb = load_workbook(buffer)
        ws = wb.active
        
        # Auto-ajustement des largeurs
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Style des en-têtes
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Sauvegarder dans un nouveau buffer
        buffer = BytesIO()
        wb.save(buffer)
    
    buffer.seek(0)
    return buffer


def create_pivot_table(
    df: pd.DataFrame,
    values: Union[str, List[str]],
    index: Union[str, List[str]],
    columns: Optional[Union[str, List[str]]] = None,
    aggfunc: Union[str, List[str], Dict] = 'sum',
    fill_value: Any = 0
) -> pd.DataFrame:
    """
    Crée un tableau croisé dynamique (pivot table) de manière sécurisée.
    
    Args:
        df: DataFrame source
        values: Colonne(s) à agréger
        index: Colonne(s) pour les lignes
        columns: Colonne(s) pour les colonnes (optionnel)
        aggfunc: Fonction d'agrégation ('sum', 'mean', 'count', 'min', 'max', etc.)
        fill_value: Valeur pour les cellules vides
    
    Returns:
        DataFrame du pivot table
    """
    try:
        pivot = pd.pivot_table(
            df,
            values=values,
            index=index,
            columns=columns,
            aggfunc=aggfunc,
            fill_value=fill_value
        )
        return pivot.reset_index()
    except Exception as e:
        raise ValueError(f"Erreur lors de la création du pivot table: {e}")


def merge_excel_files(
    file_paths_or_buffers: List[Union[str, Path, BytesIO]],
    merge_type: str = 'concat',
    merge_key: Optional[str] = None,
    how: str = 'outer'
) -> pd.DataFrame:
    """
    Fusionne plusieurs fichiers Excel.
    
    Args:
        file_paths_or_buffers: Liste des fichiers ou buffers
        merge_type: 
            - 'concat': Empile les DataFrames verticalement
            - 'merge': Joint sur une colonne commune
        merge_key: Colonne de jointure (requis si merge_type='merge')
        how: Type de jointure ('inner', 'outer', 'left', 'right')
    
    Returns:
        DataFrame fusionné
    """
    if not file_paths_or_buffers:
        raise ValueError("Aucun fichier fourni pour la fusion")
    
    dfs = []
    for file in file_paths_or_buffers:
        # Reset buffer position if needed
        if hasattr(file, 'seek'):
            file.seek(0)
        
        # Détecter le type de fichier
        if hasattr(file, 'name'):
            filename = file.name
        else:
            filename = str(file)
        
        if filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        
        dfs.append(df)
    
    if merge_type == 'concat':
        return pd.concat(dfs, ignore_index=True)
    elif merge_type == 'merge':
        if not merge_key:
            raise ValueError("merge_key requis pour merge_type='merge'")
        
        result = dfs[0]
        for df in dfs[1:]:
            result = pd.merge(result, df, on=merge_key, how=how)
        return result
    else:
        raise ValueError(f"merge_type non reconnu: {merge_type}")


def validate_excel_file(file_path_or_buffer: Union[str, Path, BytesIO]) -> Dict[str, Any]:
    """
    Valide un fichier Excel et retourne des informations.
    
    Args:
        file_path_or_buffer: Chemin du fichier ou buffer
    
    Returns:
        Dict avec informations sur le fichier:
            - valid: bool
            - sheets: list of sheet names
            - total_rows: int
            - error: str (si invalid)
    """
    try:
        sheets = detect_excel_sheets(file_path_or_buffer)
        
        total_rows = 0
        for sheet in sheets[:MAX_SHEETS_LOADED]:
            if hasattr(file_path_or_buffer, 'seek'):
                file_path_or_buffer.seek(0)
            df = pd.read_excel(file_path_or_buffer, sheet_name=sheet)
            total_rows += len(df)
        
        return {
            'valid': True,
            'sheets': sheets,
            'sheet_count': len(sheets),
            'total_rows': total_rows,
            'error': None
        }
    except Exception as e:
        return {
            'valid': False,
            'sheets': [],
            'sheet_count': 0,
            'total_rows': 0,
            'error': str(e)
        }


def should_export_to_excel(question: str, code: str, result: Any) -> bool:
    """
    Détermine si le résultat devrait être exporté en Excel.
    
    Args:
        question: Question de l'utilisateur
        code: Code généré par l'IA
        result: Résultat de l'exécution
    
    Returns:
        True si export Excel recommandé
    """
    if not isinstance(result, pd.DataFrame):
        return False
    
    excel_keywords = [
        "export", "excel", "télécharger", "sauvegarder", 
        "download", "xlsx", "exporter", "enregistrer"
    ]
    
    question_lower = question.lower()
    code_lower = code.lower()
    
    return (
        any(kw in question_lower for kw in excel_keywords) or
        "to_excel" in code_lower
    )

